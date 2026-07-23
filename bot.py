from __future__ import annotations
import asyncio
import io
import logging
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo

from aiogram import Bot, Dispatcher, F, Router
from aiogram.filters import Command, CommandStart
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import (
    BufferedInputFile,
    CallbackQuery,
    Document,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    Message,
)

from config import ADMIN_IDS, BOT_TOKEN, TIMEZONE, CALL_SERVICE_NAME
from db import (
    add_to_blacklist,
    create_call,
    get_all_attendance_for_export,
    get_all_calls_for_stats,
    get_all_scheduled_calls,
    get_attendance,
    get_attendees_for_call,
    get_available_slots_for_user,
    get_all_users,
    get_blacklist,
    get_call,
    get_funnel_by_call,
    get_funnel_by_source,
    get_funnel_stats,
    get_last_started_call,
    get_new_users_to_invite,
    get_next_call_after,
    get_attendance_manager,
    get_curators,
    get_next_manager_contact,
    get_user_curator,
    get_any_contact,
    get_setting,
    save_curators,
    set_attendance_manager,
    get_users_by_attendance_status,
    get_users_by_call_and_status,
    update_call,
    get_users_to_reinvite,
    init_db,
    invite_count,
    is_blacklisted,
    set_attendance_status,
    set_call_status,
    set_setting,
    upsert_attendance,
    upsert_user,
)
from keyboards import (
    admin_main_keyboard,
    back_keyboard,
    broadcast_confirm_keyboard,
    broadcast_scope_keyboard,
    broadcast_segment_keyboard,
    calls_keyboard,
    cancel_keyboard,
    confirmed_keyboard,
    dates_keyboard,
    home_keyboard,
    rsvp_keyboard,
    slot_keyboard,
    times_keyboard,
)
from scheduler import (
    remove_call_pings,
    restore_on_startup,
    schedule_call_pings,
    _broadcast,
)
from texts import (
    BLACKLISTED,
    CALL_CANCELED_NOTIFY,
    CALL_UPDATED_NOTIFY,
    CONTACT_ISSUED,
    INVITE_CHOOSE_SLOT,
    INVITE_NO_CALL,
    INVITE_WITH_CALL,
    NEW_SLOT_NOTIFY,
    NO_MANAGER_CONTACT,
    NO_SLOT_FIT,
    RSVP_CONFIRMED,
    RSVP_DECLINED,
    SLOT_CHOSEN,
    TRIGGER_WRONG,
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

TZ = ZoneInfo(TIMEZONE)

router = Router()
admin_router = Router()


# ══════════════════════════════════════════════════════════════════════════════
# FSM — состояния для /admin
# ══════════════════════════════════════════════════════════════════════════════

class AdminNewCall(StatesGroup):
    waiting_date = State()
    waiting_time = State()
    waiting_link = State()


class AdminVacancy(StatesGroup):
    waiting_text = State()


class AdminManager(StatesGroup):
    waiting_contact = State()


class AdminSupport(StatesGroup):
    waiting_contact = State()


class AdminCancelCall(StatesGroup):
    waiting_call_id = State()


class AdminEditCall(StatesGroup):
    waiting_call_id = State()
    waiting_field   = State()
    waiting_date    = State()
    waiting_time    = State()
    waiting_link    = State()


class AdminBroadcast(StatesGroup):
    waiting_segment = State()
    waiting_call    = State()
    waiting_text    = State()
    waiting_confirm = State()


class AdminBlacklist(StatesGroup):
    waiting_file = State()


# ══════════════════════════════════════════════════════════════════════════════
# Хэндлеры кандидата
# ══════════════════════════════════════════════════════════════════════════════

@router.message(CommandStart())
async def cmd_start(message: Message, state: FSMContext) -> None:
    user = message.from_user
    payload = message.text.split(maxsplit=1)[1] if " " in (message.text or "") else None
    source = payload if payload else None

    # Проверка чёрного списка до регистрации
    if await is_blacklisted(user.username):
        await message.answer(BLACKLISTED)
        return

    await upsert_user(
        user_id=user.id,
        username=user.username,
        full_name=user.full_name,
        source=source,
    )
    await _show_vacancy(message, user.id)


async def _show_vacancy(message: Message, user_id: int) -> None:
    vacancy_text = await get_setting("vacancy_text") or "[ВСТАВЬТЕ ОПИСАНИЕ ВАКАНСИИ]"

    slots = await get_available_slots_for_user(user_id)

    if slots:
        now = datetime.now(TZ)
        urgent = [s for s in slots if datetime.fromisoformat(s["call_dt"]) <= now + timedelta(minutes=5)]

        # Срочный созвон — показываем баннер с ссылкой перед вакансией
        if urgent:
            call = urgent[0]
            link = call.get("link", "")
            await message.answer(
                f"⚡️ Прямо сейчас идёт созвон — можете подключиться прямо сейчас:\n"
                f"<a href=\"{link}\">Войти в {CALL_SERVICE_NAME}</a>",
                parse_mode="HTML",
            )

        # Всегда показываем вакансию и все слоты (включая срочные)
        text = INVITE_CHOOSE_SLOT.format(vacancy_text=vacancy_text)
        await message.answer(text, reply_markup=slot_keyboard(slots), parse_mode="HTML")
    else:
        await message.answer(INVITE_NO_CALL.format(vacancy_text=vacancy_text), parse_mode="HTML")


@router.callback_query(F.data.startswith("slot:"))
async def handle_slot_choice(callback: CallbackQuery) -> None:
    raw = callback.data.split(":")[1]
    user_id = callback.from_user.id

    if raw == "none":
        # Контакт для подбора времени: HR задаёт сам через админку,
        # если не задан — берём ближайший доступный контакт из бота:
        # закреплённый за кандидатом куратор → любой активный куратор → запасной менеджер.
        support_contact = await get_setting("support_contact")
        if not support_contact:
            support_contact = await get_user_curator(user_id)
        if not support_contact:
            support_contact = await get_any_contact()
        if not support_contact:
            await callback.answer("Контакт пока не настроен. Попробуйте позже.", show_alert=True)
            return
        await callback.message.edit_text(
            NO_SLOT_FIT.format(support_contact=support_contact),
            parse_mode="HTML",
            reply_markup=home_keyboard(),
        )
        await callback.answer()
        return

    call_id = int(raw)
    call = await get_call(call_id)
    if not call or call["status"] != "scheduled":
        await callback.answer("Этот слот уже недоступен.", show_alert=True)
        return

    await upsert_attendance(user_id, call_id, status="invited")

    call_dt = datetime.fromisoformat(call["call_dt"])
    dt_str = call_dt.strftime("%d.%m.%Y %H:%M МСК")

    # Берём куратора: 1) уже назначен на этот созвон, 2) был назначен раньше, 3) round-robin
    manager_contact = await get_attendance_manager(user_id, call_id)
    if not manager_contact:
        manager_contact = await get_user_curator(user_id)
    if not manager_contact:
        manager_contact = await get_next_manager_contact() or "уточняется"
    await set_attendance_manager(user_id, call_id, manager_contact)

    text = SLOT_CHOSEN.format(
        call_dt=dt_str,
        link=call.get("link", ""),
        manager_contact=manager_contact,
        service=CALL_SERVICE_NAME,
    )
    await callback.message.edit_text(text, reply_markup=rsvp_keyboard(call_id), parse_mode="HTML")
    await callback.answer()



@router.callback_query(F.data.startswith("rsvp:"))
async def handle_rsvp(callback: CallbackQuery) -> None:
    parts = callback.data.split(":")
    action  = parts[1]
    call_id = int(parts[2])

    await set_attendance_status(callback.from_user.id, call_id, action)
    await callback.message.edit_reply_markup(reply_markup=None)
    await callback.answer()

    if action == "confirmed":
        await callback.message.answer(
            RSVP_CONFIRMED.format(service=CALL_SERVICE_NAME),
            reply_markup=confirmed_keyboard(call_id),
        )
    else:
        await callback.message.answer(RSVP_DECLINED)
        # Сразу показываем другие доступные слоты
        slots = await get_available_slots_for_user(callback.from_user.id)
        if slots:
            vacancy_text = await get_setting("vacancy_text") or "[ВСТАВЬТЕ ОПИСАНИЕ ВАКАНСИИ]"
            text = INVITE_CHOOSE_SLOT.format(vacancy_text=vacancy_text)
            await callback.message.answer(text, reply_markup=slot_keyboard(slots), parse_mode="HTML")
        else:
            await callback.message.answer(
                "Как только появятся новые слоты — мы сообщим вам.",
                reply_markup=home_keyboard(),
            )


@router.callback_query(F.data.startswith("cancel_slot:"))
async def handle_cancel_slot(callback: CallbackQuery) -> None:
    call_id = int(callback.data.split(":")[1])
    user_id = callback.from_user.id

    await set_attendance_status(user_id, call_id, "declined")
    await callback.message.edit_reply_markup(reply_markup=None)
    await callback.answer()

    await callback.message.answer("Понял, отменили вашу запись.")

    # Показываем доступные слоты
    slots = await get_available_slots_for_user(user_id)
    if slots:
        vacancy_text = await get_setting("vacancy_text") or "[ВСТАВЬТЕ ОПИСАНИЕ ВАКАНСИИ]"
        text = INVITE_CHOOSE_SLOT.format(vacancy_text=vacancy_text)
        await callback.message.answer(text, reply_markup=slot_keyboard(slots), parse_mode="HTML")
    else:
        await callback.message.answer(
            "Свободных слотов пока нет. Как только появятся — мы сообщим вам.",
            reply_markup=home_keyboard(),
        )


@router.callback_query(F.data == "menu:start")
async def handle_menu_start(callback: CallbackQuery) -> None:
    await callback.message.edit_reply_markup(reply_markup=None)
    await callback.answer()
    await _show_vacancy(callback.message, callback.from_user.id)


@router.message(F.text)
async def handle_text(message: Message) -> None:
    # Игнорируем команды и сообщения от админов (они обрабатываются в admin_router через FSM)
    if message.text and message.text.startswith("/"):
        return

    call = await get_last_started_call()
    if not call:
        return

    # Проверка окна активности: триггер активен до создания следующего созвона
    next_call = await get_next_call_after(call["id"])
    if next_call:
        # Если следующий созвон уже создан, триггер предыдущего неактивен
        return

    trigger = (call.get("trigger_word") or "").strip().lower()
    user_text = (message.text or "").strip().lower()

    if trigger and user_text == trigger:
        user_id = message.from_user.id
        # Берём уже назначенного куратора, не крутим счётчик заново
        manager_contact = await get_attendance_manager(user_id, call["id"])
        if not manager_contact:
            manager_contact = await get_user_curator(user_id)
        if not manager_contact:
            manager_contact = await get_next_manager_contact()
        if not manager_contact:
            await message.answer(NO_MANAGER_CONTACT)
            return
        # Найти любой attendance для этого пользователя по данному созвону
        att = await get_attendance(user_id, call["id"])
        if att and att["status"] != "contact_issued":
            await set_attendance_status(user_id, call["id"], "contact_issued")
            await set_attendance_manager(user_id, call["id"], manager_contact)
        elif not att:
            # Кандидат написал триггер, но не был привязан к созвону (зашёл позже)
            await upsert_attendance(user_id, call["id"], status="contact_issued")
            await set_attendance_manager(user_id, call["id"], manager_contact)
        await message.answer(
            CONTACT_ISSUED.format(manager_contact=manager_contact),
            parse_mode="HTML",
        )
    else:
        await message.answer(TRIGGER_WRONG)


# ══════════════════════════════════════════════════════════════════════════════
# Фильтр: только ADMIN_IDS
# ══════════════════════════════════════════════════════════════════════════════

def is_admin(user_id: int) -> bool:
    return user_id in ADMIN_IDS


# ══════════════════════════════════════════════════════════════════════════════
# /admin — точка входа
# ══════════════════════════════════════════════════════════════════════════════

@admin_router.message(Command("admin"))
async def cmd_admin(message: Message) -> None:
    if not is_admin(message.from_user.id):
        return
    await message.answer("🛠 Панель администратора:", reply_markup=admin_main_keyboard())


# ── Отмена FSM ────────────────────────────────────────────────────────────────

@admin_router.callback_query(F.data == "admin:cancel_fsm")
async def admin_cancel_fsm(callback: CallbackQuery, state: FSMContext) -> None:
    if not is_admin(callback.from_user.id):
        return
    await state.clear()
    await callback.message.edit_text("Действие отменено.")
    await callback.answer()


# ══════════════════════════════════════════════════════════════════════════════
# Назначить созвон
# ══════════════════════════════════════════════════════════════════════════════

@admin_router.callback_query(F.data == "admin:new_call")
async def admin_new_call_start(callback: CallbackQuery, state: FSMContext) -> None:
    if not is_admin(callback.from_user.id):
        return
    await state.set_state(AdminNewCall.waiting_date)
    await callback.message.answer(
        "Введите <b>дату</b> созвона в формате <b>ДД.ММ.ГГГГ</b>:",
        reply_markup=cancel_keyboard(),
        parse_mode="HTML",
    )
    await callback.answer()


@admin_router.message(AdminNewCall.waiting_date)
async def admin_new_call_date(message: Message, state: FSMContext) -> None:
    if not is_admin(message.from_user.id):
        return
    text = (message.text or "").strip()
    for fmt in ("%d.%m.%Y", "%d.%m.%y"):
        try:
            from datetime import date
            d = datetime.strptime(text, fmt).date()
            await state.update_data(call_date=text if len(text) > 8 else d.strftime("%d.%m.%Y"))
            await state.set_state(AdminNewCall.waiting_time)
            await message.answer(
                f"Дата: <b>{d.strftime('%d.%m.%Y')}</b>\n\nТеперь введите <b>время</b> созвона (МСК) в формате <b>ЧЧ:ММ</b>:",
                reply_markup=cancel_keyboard(),
                parse_mode="HTML",
            )
            return
        except ValueError:
            continue
    await message.answer(
        "Не смог распознать дату. Формат: <b>ДД.ММ.ГГГГ</b>",
        reply_markup=cancel_keyboard(),
        parse_mode="HTML",
    )


@admin_router.message(AdminNewCall.waiting_time)
async def admin_new_call_time(message: Message, state: FSMContext) -> None:
    if not is_admin(message.from_user.id):
        return
    text = (message.text or "").strip()
    data = await state.get_data()
    call_date = data.get("call_date", "")
    dt = _parse_dt(f"{call_date} {text}")
    if not dt:
        await message.answer(
            "Не смог распознать время. Формат: <b>ЧЧ:ММ</b>",
            reply_markup=cancel_keyboard(),
            parse_mode="HTML",
        )
        return
    await state.update_data(call_dt=dt.isoformat())
    await state.set_state(AdminNewCall.waiting_link)
    await message.answer("Введите ссылку на созвон:", reply_markup=cancel_keyboard())


@admin_router.message(AdminNewCall.waiting_link)
async def admin_new_call_link(message: Message, state: FSMContext, bot: Bot) -> None:
    if not is_admin(message.from_user.id):
        return
    link = (message.text or "").strip()
    if not link:
        await message.answer("Ссылка не может быть пустой.", reply_markup=cancel_keyboard())
        return
    data = await state.get_data()
    await state.update_data(link=link)
    await state.clear()

    call_dt = datetime.fromisoformat(data["call_dt"])
    call_id = await create_call(call_dt, link, trigger_word="—")
    call = await get_call(call_id)

    # Рассылка инвайтов
    max_inv = int(await get_setting("max_invites") or 3)
    vacancy_text = await get_setting("vacancy_text") or "[ВСТАВЬТЕ ОПИСАНИЕ ВАКАНСИИ]"
    dt_str = call_dt.strftime("%d.%m.%Y %H:%M МСК")

    new_users      = await get_new_users_to_invite(call_id, max_inv)
    reinvite_users = await get_users_to_reinvite(call_id, max_inv)
    all_users = new_users + reinvite_users

    # Шлём оповещение с кнопкой выбора — пусть сами выберут удобное время
    sent = 0
    if all_users:
        # Собираем все доступные слоты для показа кнопок
        all_slots = await get_all_scheduled_calls()
        kb = slot_keyboard(all_slots) if all_slots else None
        for u in all_users:
            try:
                await bot.send_message(
                    u["user_id"],
                    NEW_SLOT_NOTIFY,
                    reply_markup=kb,
                )
                sent += 1
            except Exception:
                pass

    scheduler = _get_scheduler()
    if scheduler:
        schedule_call_pings(scheduler, bot, call)

    await message.answer(
        f"✅ Созвон назначен на <b>{dt_str}</b>\n"
        f"Разослано инвайтов: <b>{sent}</b>",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="📅 Назначить ещё один созвон", callback_data="admin:new_call")],
            [InlineKeyboardButton(text="🔙 В главное меню", callback_data="admin:main")],
        ]),
    )


# ══════════════════════════════════════════════════════════════════════════════
# Редактировать вакансию
# ══════════════════════════════════════════════════════════════════════════════

@admin_router.callback_query(F.data == "admin:vacancy")
async def admin_vacancy_start(callback: CallbackQuery, state: FSMContext) -> None:
    if not is_admin(callback.from_user.id):
        return
    current = await get_setting("vacancy_text") or ""
    await state.set_state(AdminVacancy.waiting_text)
    await callback.message.answer(
        f"Текущий текст вакансии:\n\n{current}\n\nВведите новый текст:",
        reply_markup=cancel_keyboard(),
    )
    await callback.answer()


@admin_router.message(AdminVacancy.waiting_text)
async def admin_vacancy_text(message: Message, state: FSMContext) -> None:
    if not is_admin(message.from_user.id):
        return
    text = (message.text or "").strip()
    if not text:
        await message.answer("Текст не может быть пустым.", reply_markup=cancel_keyboard())
        return
    await set_setting("vacancy_text", text)
    await state.clear()
    await message.answer("✅ Текст вакансии обновлён.")


# ══════════════════════════════════════════════════════════════════════════════
# Контакт менеджера
# ══════════════════════════════════════════════════════════════════════════════

@admin_router.callback_query(F.data == "admin:manager")
async def admin_manager_start(callback: CallbackQuery, state: FSMContext) -> None:
    if not is_admin(callback.from_user.id):
        return
    current = await get_setting("manager_contact") or "(не задан)"
    await state.set_state(AdminManager.waiting_contact)
    await callback.message.answer(
        f"Текущий контакт менеджера: {current}\n\nВведите новый контакт (например, @username):",
        reply_markup=cancel_keyboard(),
        parse_mode="HTML",
    )
    await callback.answer()


@admin_router.message(AdminManager.waiting_contact)
async def admin_manager_contact(message: Message, state: FSMContext) -> None:
    if not is_admin(message.from_user.id):
        return
    contact = (message.text or "").strip()
    if not contact:
        await message.answer("Контакт не может быть пустым.", reply_markup=cancel_keyboard())
        return
    await set_setting("manager_contact", contact)
    await state.clear()
    await message.answer(f"✅ Контакт менеджера обновлён: {contact}", parse_mode="HTML")


@admin_router.callback_query(F.data == "admin:support")
async def admin_support_start(callback: CallbackQuery, state: FSMContext) -> None:
    if not is_admin(callback.from_user.id):
        return
    current = await get_setting("support_contact") or "(не задан — берётся куратор или запасной менеджер)"
    await state.set_state(AdminSupport.waiting_contact)
    await callback.message.answer(
        f"Контакт для варианта «ни одно время не подходит»: {current}\n\n"
        "Кандидат получает его, когда нажимает «Ни одно время не подходит».\n"
        "Введите новый контакт (например, @username):",
        reply_markup=cancel_keyboard(),
        parse_mode="HTML",
    )
    await callback.answer()


@admin_router.message(AdminSupport.waiting_contact)
async def admin_support_contact(message: Message, state: FSMContext) -> None:
    if not is_admin(message.from_user.id):
        return
    contact = (message.text or "").strip()
    if not contact:
        await message.answer("Контакт не может быть пустым.", reply_markup=cancel_keyboard())
        return
    await set_setting("support_contact", contact)
    await state.clear()
    await message.answer(f"✅ Контакт для подбора времени обновлён: {contact}", parse_mode="HTML")


# ══════════════════════════════════════════════════════════════════════════════
# Кураторы (round-robin)
# ══════════════════════════════════════════════════════════════════════════════

class AdminCurators(StatesGroup):
    waiting_contacts = State()


def _curators_keyboard(curators: list[dict]) -> InlineKeyboardMarkup:
    rows = []
    for i, c in enumerate(curators):
        status = "✅" if c.get("active", True) else "❌"
        toggle = "off" if c.get("active", True) else "on"
        rows.append([
            InlineKeyboardButton(text=f"{status} {c['contact']}", callback_data=f"admin:curators:toggle:{i}:{toggle}"),
        ])
    rows.append([InlineKeyboardButton(text="✏️ Изменить список", callback_data="admin:curators:edit")])
    rows.append([InlineKeyboardButton(text="🔙 Назад", callback_data="admin:main")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


async def _curators_text(curators: list[dict]) -> str:
    if not curators:
        return "👥 <b>Кураторы не настроены.</b>\n\nДобавьте через кнопку ниже."
    active = [c for c in curators if c.get("active", True)]
    fallback = await get_setting("manager_contact") or "(не задан)"
    lines = ["👥 <b>Кураторы (round-robin):</b>\n"]
    for c in curators:
        status = "✅ активен" if c.get("active", True) else "❌ отключён"
        lines.append(f"• {c['contact']} — {status}")
    lines.append(f"\nАктивных: <b>{len(active)}</b> из <b>{len(curators)}</b>")
    if not active:
        lines.append(f"\n⚠️ Все отключены — кандидаты получат: <b>{fallback}</b>")
    return "\n".join(lines)


@admin_router.callback_query(F.data == "admin:curators")
async def admin_curators(callback: CallbackQuery) -> None:
    if not is_admin(callback.from_user.id):
        return
    curators = await get_curators()
    await callback.message.answer(
        await _curators_text(curators),
        parse_mode="HTML",
        reply_markup=_curators_keyboard(curators),
    )
    await callback.answer()


@admin_router.callback_query(F.data.startswith("admin:curators:toggle:"))
async def admin_curators_toggle(callback: CallbackQuery) -> None:
    if not is_admin(callback.from_user.id):
        return
    parts = callback.data.split(":")
    idx = int(parts[3])
    action = parts[4]  # "on" или "off"
    curators = await get_curators()
    if idx >= len(curators):
        await callback.answer("Куратор не найден.")
        return
    curators[idx]["active"] = (action == "on")
    await save_curators(curators)
    await callback.message.edit_text(
        await _curators_text(curators),
        parse_mode="HTML",
        reply_markup=_curators_keyboard(curators),
    )
    name = curators[idx]["contact"]
    status_word = "включён" if action == "on" else "отключён"
    await callback.answer(f"{name} {status_word}")


@admin_router.callback_query(F.data == "admin:curators:edit")
async def admin_curators_edit(callback: CallbackQuery, state: FSMContext) -> None:
    if not is_admin(callback.from_user.id):
        return
    curators = await get_curators()
    current = ", ".join(c["contact"] for c in curators)
    await state.set_state(AdminCurators.waiting_contacts)
    await callback.message.answer(
        f"Текущий список: <code>{current or '(пусто)'}</code>\n\n"
        "Введите контакты кураторов через запятую:\n"
        "<code>@curator1, @curator2</code>",
        parse_mode="HTML",
        reply_markup=cancel_keyboard(),
    )
    await callback.answer()


@admin_router.message(AdminCurators.waiting_contacts)
async def admin_curators_save(message: Message, state: FSMContext) -> None:
    if not is_admin(message.from_user.id):
        return
    raw = (message.text or "").strip()
    contacts = [c.strip() for c in raw.split(",") if c.strip()]
    if not contacts:
        await message.answer("Введите хотя бы одного куратора.", reply_markup=cancel_keyboard())
        return
    curators = [{"contact": c, "active": True} for c in contacts]
    await save_curators(curators, reset_index=True)
    await state.clear()
    lines = ["✅ Кураторы обновлены (все активны):\n"]
    for c in curators:
        lines.append(f"• {c['contact']}")
    lines.append("\nОтключить можно через кнопки в меню кураторов.")
    await message.answer("\n".join(lines), parse_mode="HTML")


# ══════════════════════════════════════════════════════════════════════════════
# Статистика
# ══════════════════════════════════════════════════════════════════════════════


@admin_router.callback_query(F.data == "admin:stats")
async def admin_stats(callback: CallbackQuery) -> None:
    if not is_admin(callback.from_user.id):
        return

    funnel = await get_funnel_stats()

    def pct(a: int, b: int) -> str:
        return f"{round(a/b*100)}%" if b else "—"

    e  = funnel["entered"]
    inv = funnel["invited"]
    c  = funnel["confirmed"]
    d  = funnel["declined"]
    ci = funnel["contact_issued"]

    lines = [
        "📊 <b>Воронка</b>",
        f"👣 Зашли: <b>{e}</b>  📩 Инвайт: <b>{inv}</b> ({pct(inv,e)})",
        f"✅ Подтвердили: <b>{c}</b> ({pct(c,inv)})  ❌ Отказались: <b>{d}</b>",
        f"🏆 Получили контакт: <b>{ci}</b> ({pct(ci,e)})",
    ]

    by_source = await get_funnel_by_source()
    if by_source:
        lines += ["", "🔗 <b>По источникам:</b>"]
        for row in by_source:
            src = row["source"] or "прямой"
            lines.append(f"  <b>{src}</b>: {row['entered']} → {row['confirmed']} → {row['contact_issued']}")

    all_calls = await get_all_calls_for_stats()

    await callback.message.answer("\n".join(lines), parse_mode="HTML")
    if all_calls:
        await callback.message.answer(
            "Выберите дату созвона:",
            reply_markup=dates_keyboard(all_calls, back="admin:main"),
        )
    await callback.answer()


@admin_router.callback_query(F.data.startswith("stats:date:"))
async def admin_stats_date(callback: CallbackQuery) -> None:
    if not is_admin(callback.from_user.id):
        return
    date_str = callback.data.split("stats:date:")[1]  # дд.мм.гггг
    all_calls = await get_all_calls_for_stats()
    calls_for_date = [
        c for c in all_calls
        if datetime.fromisoformat(c["call_dt"]).strftime("%d.%m.%Y") == date_str
    ]
    await callback.message.edit_text(
        f"Дата: <b>{date_str}</b>\n\nВыберите время созвона:",
        parse_mode="HTML",
        reply_markup=times_keyboard(calls_for_date, back="stats:back"),
    )
    await callback.answer()


@admin_router.callback_query(F.data.startswith("stats:call:"))
async def admin_stats_call(callback: CallbackQuery) -> None:
    if not is_admin(callback.from_user.id):
        return

    call_id = int(callback.data.split(":")[2])
    call = await get_call(call_id)
    if not call:
        await callback.answer("Созвон не найден.")
        return

    dt_str = datetime.fromisoformat(call["call_dt"]).strftime("%d.%m.%Y %H:%M МСК")
    attendees = await get_attendees_for_call(call_id)

    confirmed = [u for u in attendees if u["status"] in ("confirmed", "contact_issued")]
    declined  = [u for u in attendees if u["status"] == "declined"]
    invited   = [u for u in attendees if u["status"] == "invited"]

    def fmt_user(u: dict) -> str:
        name = f"@{u['username']}" if u.get("username") else u.get("full_name") or str(u["user_id"])
        return name

    lines = [f"📅 <b>{dt_str}</b>", ""]

    lines.append(f"✅ <b>Подтвердили ({len(confirmed)}):</b>")
    lines += [f"  {fmt_user(u)}" for u in confirmed] or ["  —"]

    lines.append(f"\n⏳ <b>Не ответили ({len(invited)}):</b>")
    lines += [f"  {fmt_user(u)}" for u in invited] or ["  —"]

    lines.append(f"\n❌ <b>Отказались ({len(declined)}):</b>")
    lines += [f"  {fmt_user(u)}" for u in declined] or ["  —"]

    await callback.message.answer(
        "\n".join(lines),
        parse_mode="HTML",
        reply_markup=back_keyboard("stats:back"),
    )
    await callback.answer()


@admin_router.callback_query(F.data == "stats:back")
async def admin_stats_back(callback: CallbackQuery) -> None:
    if not is_admin(callback.from_user.id):
        return
    all_calls = await get_all_calls_for_stats()
    await callback.message.edit_text(
        "Выберите дату созвона:",
        reply_markup=dates_keyboard(all_calls, back="admin:main"),
    )
    await callback.answer()


@admin_router.callback_query(F.data == "admin:main")
async def admin_main(callback: CallbackQuery) -> None:
    if not is_admin(callback.from_user.id):
        return
    await callback.message.edit_text("🛠 Панель администратора:", reply_markup=admin_main_keyboard())
    await callback.answer()


# ══════════════════════════════════════════════════════════════════════════════
# Реферальные ссылки
# ══════════════════════════════════════════════════════════════════════════════

BOT_USERNAME = "candidatforhr_bot"

LINKS_SOURCES = [
    ("Instagram",              "instagram"),
    ("VK",                     "vk"),
    ("YouTube",                "youtube"),
    ("Telegram",               "telegram"),
    ("HH Кологривов",          "hh"),
    ("HH Петрина Элина",       "hh_petrina"),
    ("Авито",                  "avito"),
    ("Реклама",                "ads"),
]


class AdminLinks(StatesGroup):
    waiting_name = State()


@admin_router.callback_query(F.data == "admin:links")
async def admin_links(callback: CallbackQuery) -> None:
    if not is_admin(callback.from_user.id):
        return

    lines = ["🔗 <b>Реферальные ссылки</b>\n"]
    for label, source in LINKS_SOURCES:
        lines.append(f"<b>{label}:</b>\n<code>https://t.me/{BOT_USERNAME}?start={source}</code>\n")

    await callback.message.answer(
        "\n".join(lines),
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="➕ Добавить источник", callback_data="admin:links:add")],
            [InlineKeyboardButton(text="🔙 Назад", callback_data="admin:main")],
        ]),
    )
    await callback.answer()


@admin_router.callback_query(F.data == "admin:links:add")
async def admin_links_add(callback: CallbackQuery, state: FSMContext) -> None:
    if not is_admin(callback.from_user.id):
        return
    await state.set_state(AdminLinks.waiting_name)
    await callback.message.answer(
        "Введите название источника (например: <b>TikTok</b>).\n"
        "Ссылка будет создана автоматически.",
        parse_mode="HTML",
        reply_markup=cancel_keyboard(),
    )
    await callback.answer()


@admin_router.message(AdminLinks.waiting_name)
async def admin_links_add_name(message: Message, state: FSMContext) -> None:
    if not is_admin(message.from_user.id):
        return
    name = (message.text or "").strip()
    if not name:
        await message.answer("Название не может быть пустым.", reply_markup=cancel_keyboard())
        return
    source = name.lower().replace(" ", "_")
    await state.clear()
    await message.answer(
        f"✅ Ссылка для источника <b>{name}</b>:\n\n"
        f"<code>https://t.me/{BOT_USERNAME}?start={source}</code>",
        parse_mode="HTML",
        reply_markup=back_keyboard("admin:links"),
    )


# ══════════════════════════════════════════════════════════════════════════════
# Экспорт CSV
# ══════════════════════════════════════════════════════════════════════════════

@admin_router.callback_query(F.data == "admin:export")
async def admin_export(callback: CallbackQuery) -> None:
    if not is_admin(callback.from_user.id):
        return
    rows = await get_all_attendance_for_export()
    bl = set(await get_blacklist())
    rows = [r for r in rows if (r.get("username") or "").lstrip("@").strip().lower() not in bl]

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    STATUS_RU = {
        "invited":        "Приглашён",
        "confirmed":      "Подтвердил",
        "declined":       "Отказался",
        "contact_issued": "Получил контакт",
    }
    STATUS_COLOR = {
        "invited":        "FFF9C4",  # жёлтый
        "confirmed":      "C8E6C9",  # зелёный
        "declined":       "FFCDD2",  # красный
        "contact_issued": "BBDEFB",  # синий
    }

    HEADERS = [
        "Telegram ID", "Username", "ФИО", "Источник", "Дата входа",
        "Созвон", "Статус", "Дата инвайта",
        "Дата RSVP", "Дата выдачи контакта",
    ]
    COL_WIDTHS = [14, 16, 28, 16, 18, 18, 18, 18, 18, 22]

    wb = Workbook()
    ws = wb.active
    ws.title = "Кандидаты"

    # Заголовки
    header_font    = Font(bold=True, color="FFFFFF", size=11)
    header_fill    = PatternFill("solid", fgColor="1565C0")
    header_align   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin           = Side(style="thin", color="BDBDBD")
    cell_border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = cell_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    # Данные
    data_align  = Alignment(vertical="center", wrap_text=False)
    center_align = Alignment(horizontal="center", vertical="center")

    for row_idx, row in enumerate(rows, start=2):
        status_key = row.get("attendance_status") or ""
        row_fill   = PatternFill("solid", fgColor=STATUS_COLOR.get(status_key, "FFFFFF"))

        values = [
            row.get("user_id", ""),
            f"@{row['username']}" if row.get("username") else "",
            row.get("full_name", ""),
            row.get("source") or "прямой вход",
            _fmt_dt(row.get("entered_at")),
            _fmt_dt(row.get("call_dt")) if row.get("call_dt") else "",
            STATUS_RU.get(status_key, status_key),
            _fmt_dt(row.get("invited_at")),
            _fmt_dt(row.get("rsvp_at")),
            _fmt_dt(row.get("contact_issued_at")),
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border    = cell_border
            cell.fill      = row_fill
            cell.alignment = center_align if col_idx in (1, 6) else data_align

        ws.row_dimensions[row_idx].height = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    now_str = datetime.now(TZ).strftime("%d.%m.%Y_%H-%M")
    file = BufferedInputFile(buf.read(), filename=f"candidates_{now_str}.xlsx")
    await callback.message.answer_document(file, caption=f"📊 Экспорт кандидатов — {now_str}")
    await callback.answer()


# ══════════════════════════════════════════════════════════════════════════════
# Редактировать созвон
# ══════════════════════════════════════════════════════════════════════════════

@admin_router.callback_query(F.data == "admin:edit_call")
async def admin_edit_call_start(callback: CallbackQuery, state: FSMContext) -> None:
    if not is_admin(callback.from_user.id):
        return
    calls = await get_all_scheduled_calls()
    if not calls:
        await callback.message.answer("Нет запланированных созвонов.")
        await callback.answer()
        return
    await state.set_state(AdminEditCall.waiting_call_id)
    await callback.message.answer(
        "Выберите созвон для редактирования:",
        reply_markup=calls_keyboard(calls, prefix="editcall", back="admin:main"),
    )
    await callback.answer()


@admin_router.callback_query(F.data.startswith("editcall:"), AdminEditCall.waiting_call_id)
async def admin_edit_call_chosen(callback: CallbackQuery, state: FSMContext) -> None:
    if not is_admin(callback.from_user.id):
        return
    call_id = int(callback.data.split(":")[1])
    call = await get_call(call_id)
    if not call:
        await callback.answer("Созвон не найден.")
        return
    dt_str = datetime.fromisoformat(call["call_dt"]).strftime("%d.%m.%Y %H:%M МСК")
    await state.update_data(edit_call_id=call_id, edit_call_dt=call["call_dt"], edit_link=call["link"])
    await state.set_state(AdminEditCall.waiting_field)
    await callback.message.edit_text(
        f"Созвон: <b>{dt_str}</b>\n"
        f"Ссылка: {call['link']}\n\n"
        "Что хотите изменить?",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="📅 Дату и время", callback_data="editfield:dt")],
            [InlineKeyboardButton(text="🔗 Ссылку",       callback_data="editfield:link")],
            [InlineKeyboardButton(text="📅+🔗 Всё сразу", callback_data="editfield:both")],
            [InlineKeyboardButton(text="🔙 Назад",        callback_data="admin:cancel_fsm")],
        ]),
    )
    await callback.answer()


@admin_router.callback_query(F.data.startswith("editfield:"), AdminEditCall.waiting_field)
async def admin_edit_field(callback: CallbackQuery, state: FSMContext) -> None:
    if not is_admin(callback.from_user.id):
        return
    field = callback.data.split(":")[1]
    await state.update_data(edit_field=field)
    if field in ("dt", "both"):
        await state.set_state(AdminEditCall.waiting_date)
        await callback.message.answer(
            "Введите новую <b>дату</b> (ДД.ММ.ГГГГ):",
            parse_mode="HTML",
            reply_markup=cancel_keyboard(),
        )
    else:
        await state.set_state(AdminEditCall.waiting_link)
        await callback.message.answer("Введите новую ссылку на созвон:", reply_markup=cancel_keyboard())
    await callback.answer()


@admin_router.message(AdminEditCall.waiting_date)
async def admin_edit_date(message: Message, state: FSMContext) -> None:
    if not is_admin(message.from_user.id):
        return
    text = (message.text or "").strip()
    for fmt in ("%d.%m.%Y", "%d.%m.%y"):
        try:
            from datetime import date
            d = datetime.strptime(text, fmt).date()
            await state.update_data(edit_new_date=d.strftime("%d.%m.%Y"))
            await state.set_state(AdminEditCall.waiting_time)
            await message.answer(
                f"Дата: <b>{d.strftime('%d.%m.%Y')}</b>\n\nВведите новое <b>время</b> (ЧЧ:ММ):",
                parse_mode="HTML",
                reply_markup=cancel_keyboard(),
            )
            return
        except ValueError:
            continue
    await message.answer("Не смог распознать дату. Формат: <b>ДД.ММ.ГГГГ</b>", parse_mode="HTML", reply_markup=cancel_keyboard())


@admin_router.message(AdminEditCall.waiting_time)
async def admin_edit_time(message: Message, state: FSMContext) -> None:
    if not is_admin(message.from_user.id):
        return
    data = await state.get_data()
    dt = _parse_dt(f"{data['edit_new_date']} {(message.text or '').strip()}")
    if not dt:
        await message.answer("Не смог распознать время. Формат: <b>ЧЧ:ММ</b>", parse_mode="HTML", reply_markup=cancel_keyboard())
        return
    await state.update_data(edit_new_dt=dt.isoformat())
    field = data.get("edit_field")
    if field == "both":
        await state.set_state(AdminEditCall.waiting_link)
        await message.answer("Теперь введите новую ссылку на созвон:", reply_markup=cancel_keyboard())
    else:
        await _apply_edit(message, state)


@admin_router.message(AdminEditCall.waiting_link)
async def admin_edit_link(message: Message, state: FSMContext) -> None:
    if not is_admin(message.from_user.id):
        return
    link = (message.text or "").strip()
    if not link:
        await message.answer("Ссылка не может быть пустой.", reply_markup=cancel_keyboard())
        return
    await state.update_data(edit_new_link=link)
    await _apply_edit(message, state)


async def _apply_edit(message: Message, state: FSMContext) -> None:
    """Применяет изменения созвона и рассылает уведомления участникам."""
    data = await state.get_data()
    await state.clear()

    call_id   = data["edit_call_id"]
    field     = data["edit_field"]
    new_dt    = datetime.fromisoformat(data["edit_new_dt"]) if data.get("edit_new_dt") else None
    new_link  = data.get("edit_new_link")

    # Если ссылка не менялась — берём старую
    if not new_link:
        new_link = data.get("edit_link")
    # Если дата не менялась — берём старую
    if not new_dt:
        new_dt = datetime.fromisoformat(data["edit_call_dt"])

    await update_call(call_id, call_dt=new_dt if field in ("dt", "both") else None,
                      link=new_link if field in ("link", "both") else None)

    call = await get_call(call_id)
    dt_str = datetime.fromisoformat(call["call_dt"]).strftime("%d.%m.%Y %H:%M МСК")

    # Перепланируем пуши
    scheduler = _get_scheduler()
    if scheduler:
        remove_call_pings(scheduler, call_id)
        schedule_call_pings(scheduler, message.bot, call)

    # Оповещаем всех invited + confirmed
    attendees = await get_attendees_for_call(call_id)
    notify_ids = [u["user_id"] for u in attendees if u["status"] in ("invited", "confirmed")]

    await message.answer(
        f"✅ Созвон обновлён: <b>{dt_str}</b>\n"
        f"Ссылка: {call['link']}",
        parse_mode="HTML",
    )

    if notify_ids:
        text = CALL_UPDATED_NOTIFY.format(call_dt=dt_str, link=call["link"])
        sent, _ = await _broadcast(message.bot, notify_ids, text)
        await message.answer(f"📢 Уведомлено участников: <b>{sent}</b>", parse_mode="HTML")


# ══════════════════════════════════════════════════════════════════════════════
# Отменить созвон
# ══════════════════════════════════════════════════════════════════════════════

@admin_router.callback_query(F.data == "admin:cancel_call")
async def admin_cancel_call_start(callback: CallbackQuery, state: FSMContext) -> None:
    if not is_admin(callback.from_user.id):
        return
    calls = await get_all_scheduled_calls()
    if not calls:
        await callback.message.answer("Нет запланированных созвонов.")
        await callback.answer()
        return

    lines = ["Выберите созвон для отмены (введите ID):"]
    for c in calls:
        dt_str = datetime.fromisoformat(c["call_dt"]).strftime("%d.%m.%Y %H:%M")
        lines.append(f"  <b>{c['id']}</b> — {dt_str}")

    await state.set_state(AdminCancelCall.waiting_call_id)
    await callback.message.answer("\n".join(lines), reply_markup=cancel_keyboard(), parse_mode="HTML")
    await callback.answer()


@admin_router.message(AdminCancelCall.waiting_call_id)
async def admin_cancel_call_id(message: Message, state: FSMContext) -> None:
    if not is_admin(message.from_user.id):
        return
    text = (message.text or "").strip()
    if not text.isdigit():
        await message.answer("Введите числовой ID созвона.", reply_markup=cancel_keyboard())
        return
    call_id = int(text)
    call = await get_call(call_id)
    if not call or call["status"] != "scheduled":
        await message.answer("Созвон не найден или уже не активен.", reply_markup=cancel_keyboard())
        return

    # Собираем всех записавшихся до отмены
    attendees = await get_attendees_for_call(call_id)
    notify_ids = [
        u["user_id"] for u in attendees
        if u["status"] in ("invited", "confirmed")
    ]

    await set_call_status(call_id, "canceled")
    scheduler = _get_scheduler()
    if scheduler:
        remove_call_pings(scheduler, call_id)

    await state.clear()
    dt_str = datetime.fromisoformat(call["call_dt"]).strftime("%d.%m.%Y %H:%M")
    await message.answer(f"✅ Созвон {dt_str} отменён, пуши сняты.")

    # Оповещаем записавшихся
    if notify_ids:
        bot = message.bot
        text = CALL_CANCELED_NOTIFY.format(call_dt=dt_str)
        sent, _ = await _broadcast(bot, notify_ids, text)
        await message.answer(f"📢 Оповещено участников: <b>{sent}</b>", parse_mode="HTML")


# ══════════════════════════════════════════════════════════════════════════════
# Чёрный список
# ══════════════════════════════════════════════════════════════════════════════

@admin_router.callback_query(F.data == "admin:blacklist")
async def admin_blacklist_start(callback: CallbackQuery, state: FSMContext) -> None:
    if not is_admin(callback.from_user.id):
        return
    bl = await get_blacklist()
    count_line = f"Сейчас в ЧС: <b>{len(bl)}</b> username(-ов).\n\n" if bl else "ЧС пока пуст.\n\n"
    await state.set_state(AdminBlacklist.waiting_file)
    await callback.message.answer(
        f"🚫 <b>Чёрный список</b>\n\n"
        f"{count_line}"
        f"Отправьте <b>.xlsx</b>-файл с колонкой <code>username</code> "
        f"(те же username, что в экспорте бота, без @).\n"
        f"Новые записи добавятся к существующим.",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="📋 Выгрузить все username из бота", callback_data="admin:export_usernames")],
            [InlineKeyboardButton(text="🔙 Отмена", callback_data="admin:cancel_fsm")],
        ]),
        parse_mode="HTML",
    )
    await callback.answer()


@admin_router.callback_query(F.data == "admin:export_usernames")
async def admin_export_usernames(callback: CallbackQuery) -> None:
    if not is_admin(callback.from_user.id):
        return

    users = await get_all_users()
    usernames = [u["username"] for u in users if u.get("username")]

    if not usernames:
        await callback.answer("Нет пользователей с username.", show_alert=True)
        return

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Usernames"
    ws.cell(row=1, column=1, value="username")
    for i, uname in enumerate(usernames, start=2):
        ws.cell(row=i, column=1, value=uname)
    ws.column_dimensions["A"].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    now_str = datetime.now(TZ).strftime("%d.%m.%Y_%H-%M")
    file = BufferedInputFile(buf.read(), filename=f"usernames_{now_str}.xlsx")
    await callback.message.answer_document(
        file,
        caption=f"📋 Все username из бота — {len(usernames)} шт.\nМожно загрузить в ЧС через кнопку выше.",
    )
    await callback.answer()


@admin_router.message(AdminBlacklist.waiting_file, F.document)
async def admin_blacklist_file(message: Message, state: FSMContext, bot: Bot) -> None:
    if not is_admin(message.from_user.id):
        return
    doc: Document = message.document
    if not doc.file_name or not doc.file_name.lower().endswith(".xlsx"):
        await message.answer("Нужен файл формата <b>.xlsx</b>.", parse_mode="HTML", reply_markup=cancel_keyboard())
        return

    file = await bot.get_file(doc.file_id)
    buf = io.BytesIO()
    await bot.download_file(file.file_path, destination=buf)
    buf.seek(0)

    try:
        from openpyxl import load_workbook
        wb = load_workbook(buf, read_only=True, data_only=True)
        ws = wb.active

        # Ищем колонку username (первая строка — заголовки)
        headers = [str(cell.value).strip().lower() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        if "username" not in headers:
            await message.answer(
                "В файле не найдена колонка <code>username</code>. Проверьте заголовки.",
                parse_mode="HTML",
                reply_markup=cancel_keyboard(),
            )
            return

        col_idx = headers.index("username")
        usernames = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            val = row[col_idx]
            if val:
                usernames.append(str(val).strip())
        wb.close()
    except Exception as e:
        logger.error("Blacklist file parse error: %s", e)
        await message.answer("Не удалось прочитать файл. Убедитесь, что это корректный .xlsx.", reply_markup=cancel_keyboard())
        return

    if not usernames:
        await message.answer("В файле не найдено ни одного username.", reply_markup=cancel_keyboard())
        return

    added = await add_to_blacklist(usernames)
    await state.clear()
    await message.answer(
        f"✅ Готово!\n"
        f"Обработано строк: <b>{len(usernames)}</b>\n"
        f"Добавлено новых в ЧС: <b>{added}</b>\n"
        f"(уже существующие пропущены)",
        parse_mode="HTML",
    )


@admin_router.message(AdminBlacklist.waiting_file)
async def admin_blacklist_wrong(message: Message) -> None:
    if not is_admin(message.from_user.id):
        return
    await message.answer("Пожалуйста, отправьте файл <b>.xlsx</b>.", parse_mode="HTML", reply_markup=cancel_keyboard())


# ══════════════════════════════════════════════════════════════════════════════
# Рассылка
# ══════════════════════════════════════════════════════════════════════════════

_SEGMENT_LABELS = {
    "all":           "Все кандидаты (ввели ФИО)",
    "invited":       "Не ответили на инвайт",
    "confirmed":     "Подтвердили участие",
    "declined":      "Отказались",
    "contact_issued": "Получили контакт",
}


@admin_router.callback_query(F.data == "admin:broadcast")
async def admin_broadcast_start(callback: CallbackQuery, state: FSMContext) -> None:
    if not is_admin(callback.from_user.id):
        return
    await state.set_state(AdminBroadcast.waiting_segment)
    await callback.message.answer(
        "📢 <b>Рассылка</b>\n\nВыберите сегмент получателей:",
        reply_markup=broadcast_segment_keyboard(),
        parse_mode="HTML",
    )
    await callback.answer()


@admin_router.callback_query(F.data.startswith("admin:broadcast:") & ~F.data.in_({"admin:broadcast:send"}))
async def admin_broadcast_segment(callback: CallbackQuery, state: FSMContext) -> None:
    if not is_admin(callback.from_user.id):
        return
    segment = callback.data.split("admin:broadcast:")[1]
    if segment not in _SEGMENT_LABELS:
        await callback.answer("Неизвестный сегмент.")
        return
    await state.update_data(segment=segment)
    await state.set_state(AdminBroadcast.waiting_call)
    await callback.message.answer(
        f"Сегмент: <b>{_SEGMENT_LABELS[segment]}</b>\n\n"
        "Выберите охват рассылки:",
        reply_markup=broadcast_scope_keyboard(),
        parse_mode="HTML",
    )
    await callback.answer()


@admin_router.callback_query(F.data == "bscope:all", AdminBroadcast.waiting_call)
async def admin_broadcast_scope_all(callback: CallbackQuery, state: FSMContext) -> None:
    if not is_admin(callback.from_user.id):
        return
    await state.update_data(call_id=None)
    await state.set_state(AdminBroadcast.waiting_text)
    await callback.message.answer(
        "Введите текст сообщения для рассылки:",
        reply_markup=cancel_keyboard(),
    )
    await callback.answer()


@admin_router.callback_query(F.data == "bscope:call", AdminBroadcast.waiting_call)
async def admin_broadcast_scope_call(callback: CallbackQuery, state: FSMContext) -> None:
    if not is_admin(callback.from_user.id):
        return
    all_calls = await get_all_calls_for_stats()
    if not all_calls:
        await callback.answer("Нет созвонов в базе.", show_alert=True)
        return
    await callback.message.answer(
        "Выберите созвон:",
        reply_markup=calls_keyboard(all_calls, prefix="bscall", back="admin:cancel_fsm"),
    )
    await callback.answer()


@admin_router.callback_query(F.data.startswith("bscall:"), AdminBroadcast.waiting_call)
async def admin_broadcast_call_chosen(callback: CallbackQuery, state: FSMContext) -> None:
    if not is_admin(callback.from_user.id):
        return
    call_id = int(callback.data.split(":")[1])
    call = await get_call(call_id)
    if not call:
        await callback.answer("Созвон не найден.")
        return
    dt_str = datetime.fromisoformat(call["call_dt"]).strftime("%d.%m.%Y %H:%M МСК")
    await state.update_data(call_id=call_id, call_dt_str=dt_str)
    await state.set_state(AdminBroadcast.waiting_text)
    await callback.message.answer(
        f"Созвон: <b>{dt_str}</b>\n\nВведите текст сообщения:",
        reply_markup=cancel_keyboard(),
        parse_mode="HTML",
    )
    await callback.answer()


@admin_router.message(AdminBroadcast.waiting_text)
async def admin_broadcast_text(message: Message, state: FSMContext) -> None:
    if not is_admin(message.from_user.id):
        return
    text = (message.text or "").strip()
    if not text:
        await message.answer("Текст не может быть пустым.", reply_markup=cancel_keyboard())
        return

    data = await state.get_data()
    segment = data["segment"]
    call_id = data.get("call_id")
    call_dt_str = data.get("call_dt_str")

    if call_id:
        users = await get_users_by_call_and_status(call_id, segment)
        scope_label = f"Созвон {call_dt_str}"
    else:
        users = await get_users_by_attendance_status(segment)
        scope_label = "Все кандидаты"

    count = len(users)
    await state.update_data(broadcast_text=text)
    await state.set_state(AdminBroadcast.waiting_confirm)

    preview = text if len(text) <= 200 else text[:200] + "…"
    await message.answer(
        f"📢 <b>Превью рассылки</b>\n\n"
        f"Охват: <b>{scope_label}</b>\n"
        f"Сегмент: <b>{_SEGMENT_LABELS[segment]}</b>\n"
        f"Получателей: <b>{count}</b>\n\n"
        f"Текст:\n<blockquote>{preview}</blockquote>\n\n"
        f"Отправить?",
        reply_markup=broadcast_confirm_keyboard(),
        parse_mode="HTML",
    )


@admin_router.callback_query(F.data == "admin:broadcast:send")
async def admin_broadcast_send(callback: CallbackQuery, state: FSMContext, bot: Bot) -> None:
    if not is_admin(callback.from_user.id):
        return
    data = await state.get_data()
    segment = data.get("segment")
    text = data.get("broadcast_text")
    call_id = data.get("call_id")

    if not segment or not text:
        await callback.answer("Ошибка: данные рассылки не найдены.")
        await state.clear()
        return

    await state.clear()
    await callback.message.edit_reply_markup(reply_markup=None)
    await callback.answer()

    if call_id:
        users = await get_users_by_call_and_status(call_id, segment)
    else:
        users = await get_users_by_attendance_status(segment)

    user_ids = [u["user_id"] for u in users]

    await callback.message.answer(
        f"⏳ Начинаю рассылку для <b>{len(user_ids)}</b> получателей…",
        parse_mode="HTML",
    )

    sent, failed = await _broadcast(bot, user_ids, text)

    await callback.message.answer(
        f"✅ Рассылка завершена.\n"
        f"Отправлено: <b>{sent}</b> / {len(user_ids)}\n"
        f"Не доставлено: <b>{failed}</b>",
        parse_mode="HTML",
    )


# ══════════════════════════════════════════════════════════════════════════════
# Утилиты
# ══════════════════════════════════════════════════════════════════════════════

def _fmt_dt(iso: str | None) -> str:
    if not iso:
        return ""
    try:
        return datetime.fromisoformat(iso).strftime("%d.%m.%Y %H:%M")
    except ValueError:
        return iso


def _parse_dt(text: str) -> datetime | None:
    """Парсит дату в формате ДД.ММ.ГГГГ ЧЧ:ММ (МСК)."""
    for fmt in ("%d.%m.%Y %H:%M", "%d.%m.%y %H:%M"):
        try:
            dt = datetime.strptime(text.strip(), fmt)
            return dt.replace(tzinfo=TZ)
        except ValueError:
            continue
    return None


_scheduler_ref: AsyncIOScheduler | None = None


def _get_scheduler() -> AsyncIOScheduler | None:
    return _scheduler_ref


# ══════════════════════════════════════════════════════════════════════════════
# Точка входа
# ══════════════════════════════════════════════════════════════════════════════

async def main() -> None:
    global _scheduler_ref

    await init_db()

    bot = Bot(token=BOT_TOKEN)

    from apscheduler.schedulers.asyncio import AsyncIOScheduler
    scheduler = AsyncIOScheduler(timezone=TZ)
    _scheduler_ref = scheduler
    scheduler.start()

    await restore_on_startup(scheduler, bot)

    dp = Dispatcher(storage=MemoryStorage())
    # admin_router регистрируется первым, чтобы FSM-хэндлеры имели приоритет
    dp.include_router(admin_router)
    dp.include_router(router)

    logger.info("Bot started.")
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
